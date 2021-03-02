import sys

from PyQt5.QtCore import Qt, QMimeData, QDate
from PyQt5.QtGui import QDrag
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QPushButton, QWidget, QApplication, QGridLayout, QScrollArea, QMainWindow, QSlider, QLabel, QFileDialog, QInputDialog
import math
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from mysql.connector import MySQLConnection, Error

#Otvori vezu sa bazom
vezaSaBazom = MySQLConnection(host='localhost', port="3333",
                               database='emd',
                               user='root', password='1234')

vezaSaBazom.autocommit = True

#Incijalizacija cursor objekta
cursor = vezaSaBazom.cursor()

class Stroj:
    
    def __init__(self, rok, naziv, trajanje):
        self.rok = rok
        self.naziv = naziv
        self.trajanje = trajanje
        

class Button(QPushButton):

    def __init__(self, title, parent):
        super().__init__(title, parent)

    def mouseMoveEvent(self, e):

        if e.buttons() != Qt.LeftButton:
            return

        mimeData = QMimeData()
        mimeData.setText(self.text())

        self.drag = QDrag(self)
        self.drag.setMimeData(mimeData)
        self.drag.setPixmap(self.grab())
        self.drag.setHotSpot(self.rect().center())

        dropAction = self.drag.exec_(Qt.MoveAction)
        
    def mousePressEvent(self, e):

        super().mousePressEvent(e)

        if e.button() == Qt.RightButton:
            text, result = QInputDialog.getText(self, 'Input Dialog', 'Upiši zadatak:')
            if(result == True):
                self.setText(text)


class MainWindow(QMainWindow):
    
    def izvoz(self):
        wb = Workbook()
        ws = wb.active
        today = QDate.currentDate()
        
        for x in range(self.layout.rowCount() - 1):
            for y in range(self.layout.columnCount()):
                check = type(self.layout.itemAtPosition(0, 0))
                tp = type(self.layout.itemAtPosition(x, y))
                if(tp == check):
                    continue
                else:
                    txt = self.layout.itemAtPosition(x, y).widget().text()
                    d = ws.cell(row=x + 1, column=y + 1, value=txt)
                    
        for i in range(8):
            ws.column_dimensions[get_column_letter(i+1)].width = 25
        
        wb.save("Raspored "+today.toString()+".xlsx")
    
    def uvoz(self):
        fileName = QFileDialog.getOpenFileName(self, "Odaberi postojeći raspored", "/", "CSV Files (*.csv)")
        data = []
        for x in range(len(self.btns)):
            data.append(0)
        
        with open(fileName[0], newline='', encoding='utf-8') as csvfile:
            data = csv.reader(csvfile)
            for row in data:
                new = ""
                for x in row:
                    new += x
                self.btns[self.j].setText(new)
                self.j += 1
    
    def toCSV(self):
        txt = []
        today = QDate().currentDate()
        file = open('Raspored_'+today.toString()+'.csv', 'w+', newline ='', encoding="utf-8")
        for x in range(len(self.btns)):
            txt.append(0)
        
        for x in range(len(self.btns)):
            txt[x] = self.btns[x].text()
            
        with file:     
            write = csv.writer(file) 
            write.writerows(txt) 
        
        file.close()
    
    def ubaci(self, stroj, minimum, date, row):
        for x in range(len(stroj)):
            date = date.fromString(str(stroj[x].rok), "yyyy-MM-dd")
            #date = date.addDays(-stroj[x].trajanje)
            tmp1 = date.weekNumber()
            tmp2 = tmp1[0]
            tmp = date.dayOfWeek()
            self.smjesti(0, row, stroj[x].trajanje, tmp2, minimum, tmp, stroj[x].naziv)
        return
    
    def smjesti(self, x, row, dana, tmp2, minimum, tmp, naziv):
        if(x == dana and self.layout.itemAtPosition(row + (tmp2 - minimum) * 5, tmp).widget().text() == "/"):
            for z in range(dana):
                if((tmp + z) > 7):
                    tmp = 1 - z
                    tmp2 = tmp2 + 1
                self.layout.itemAtPosition(row + (tmp2 - minimum) * 5, tmp + z).widget().setText(str(naziv))
                
            return
            
        if(self.layout.itemAtPosition(row + (tmp2 - minimum) * 5, tmp).widget().text() == "/"):
            if(tmp == 1):
                tmp = 8
                tmp2 = tmp2 - 1
                
            if(tmp2 < minimum):
                print("Zadatak "+str(naziv)+" ne stane!")
                return
            return self.smjesti(x + 1, row, dana, tmp2, minimum, tmp - 1, naziv)
        else:
            if(tmp == 1):
                tmp = 8
                tmp2 -= 1
                
            if(tmp2 < minimum):
                print("Zadatak "+str(naziv)+" ne stane!")
                return
            return self.smjesti(0, row, dana, tmp2, minimum, tmp - 1, naziv)
            

    def __init__(self):
        super().__init__()

        self.initUI()
    
        
    def initUI(self):

        
        self.layout = QGridLayout()
        self.setLayout(self.layout)
        self.btns = []
        self.spremi = QPushButton("Save")
        self.uvezi = QPushButton("Import")
        self.prebaci = QPushButton("Export to Excell")
        self.i = 0
        self.j = 0

        self.setAcceptDrops(True)
        self.scroll = QScrollArea()
        self.widget = QWidget()
        self.drag = QDrag(self)
        
        SL = []
        ST1 = []
        ST2 = []
        ostalo = []
        date = QDate()
        label = []
        maximum = 0
        minimum = 53
        
        cursor.execute('SELECT naziv, dana, rok FROM pozicija JOIN pozicijaNarudzba ON pozicija.nacrt = pozicijaNarudzba.nacrt JOIN narudzba ON pozicijaNarudzba.narudzbenica = narudzba.narudzbenica WHERE pozicija.nacrt IN (SELECT nacrt FROM tehnologijapozicija WHERE cnc LIKE "SL")')
        tmp = cursor.fetchall()
        
        for x in range(len(tmp)):
            SL.append(x)
        
        for x in range(len(tmp)):
            SL[x] = Stroj(tmp[x][2], tmp[x][0], tmp[x][1])
            
        
        cursor.execute('SELECT naziv, dana, rok FROM pozicija JOIN pozicijaNarudzba ON pozicija.nacrt = pozicijaNarudzba.nacrt JOIN narudzba ON pozicijaNarudzba.narudzbenica = narudzba.narudzbenica WHERE pozicija.nacrt IN (SELECT nacrt FROM tehnologijapozicija WHERE cnc LIKE "ST1")')
        tmp = cursor.fetchall()
        
        for x in range(len(tmp)):
            ST1.append(x)
        
        for x in range(len(tmp)):
            ST1[x] = Stroj(tmp[x][2], tmp[x][0], tmp[x][1])
        
        cursor.execute('SELECT naziv, dana, rok FROM pozicija JOIN pozicijaNarudzba ON pozicija.nacrt = pozicijaNarudzba.nacrt JOIN narudzba ON pozicijaNarudzba.narudzbenica = narudzba.narudzbenica WHERE pozicija.nacrt IN (SELECT nacrt FROM tehnologijapozicija WHERE cnc LIKE "ST2")')
        tmp = cursor.fetchall()
        
        for x in range(len(tmp)):
            ST2.append(x)
        
        for x in range(len(tmp)):
            ST2[x] = Stroj(tmp[x][2], tmp[x][0], tmp[x][1])
        
        cursor.execute('SELECT naziv, dana, rok FROM pozicija JOIN pozicijaNarudzba ON pozicija.nacrt = pozicijaNarudzba.nacrt JOIN narudzba ON pozicijaNarudzba.narudzbenica = narudzba.narudzbenica WHERE pozicija.nacrt IN (SELECT nacrt FROM tehnologijapozicija WHERE NOT cnc LIKE "SL" AND NOT cnc LIKE "ST1" AND NOT cnc LIKE "ST2")')
        tmp = cursor.fetchall()
        
        for x in range(len(tmp)):
            ostalo.append(x)
        
        for x in range(len(tmp)):
            ostalo[x] = Stroj(tmp[x][2], tmp[x][0], tmp[x][1])
        
        for x in range(7):
            label.append(0)
            
        label = (QLabel("Ponedjeljak"), QLabel("Utorak"), QLabel("Srijeda"), QLabel("Četvrtak"), QLabel("Petak"), QLabel("Subota"), QLabel("Nedjelja"))
        
        for x in range(7):
            self.layout.addWidget(label[x], 0, x+1)
            
        for x in range(len(SL)):
            date = date.fromString(str(SL[x].rok), "yyyy-MM-dd")
            SL[x].trajanje = math.ceil(SL[x].trajanje)
            date = date.addDays(-SL[x].trajanje)
            tmp = date.weekNumber()
            if int(str(tmp[0])) > maximum:
                maximum = tmp[0]
            if int(str(tmp[0])) < minimum:
                minimum = tmp[0]
                
        for x in range(len(ST1)):
            date = date.fromString(str(ST1[x].rok), "yyyy-MM-dd")
            ST1[x].trajanje = math.ceil(ST1[x].trajanje)
            date = date.addDays(-ST1[x].trajanje)
            tmp = date.weekNumber()
            if int(str(tmp[0])) > maximum:
                maximum = tmp[0]
            if int(str(tmp[0])) < minimum:
                minimum = tmp[0]
                
        for x in range(len(ST2)):
            date = date.fromString(str(ST2[x].rok), "yyyy-MM-dd")
            ST2[x].trajanje = math.ceil(ST2[x].trajanje)
            date = date.addDays(-ST2[x].trajanje)
            tmp = date.weekNumber()
            if int(str(tmp[0])) > maximum:
                maximum = tmp[0]
            if int(str(tmp[0])) < minimum:
                minimum = tmp[0]
                
        for x in range(len(ostalo)):
            date = date.fromString(str(ostalo[x].rok), "yyyy-MM-dd")
            ostalo[x].trajanje = math.ceil(ostalo[x].trajanje)
            date = date.addDays(-ostalo[x].trajanje)
            tmp = date.weekNumber()
            if int(str(tmp[0])) > maximum:
                maximum = tmp[0]
            if int(str(tmp[0])) < minimum:
                minimum = tmp[0]
                
        for x in range((maximum - minimum + 1) * 28):
            self.btns.append(0)
        
        for x in range(maximum - minimum + 1):
            label = QLabel(str(minimum + x)+". Tjedan")
            self.layout.addWidget(label, 1 + x * 5, 0)
            label = (QLabel("SL"), QLabel("ST1"), QLabel("ST2"), QLabel("Ostalo"))
            for y in range(4):
                self.layout.addWidget(label[y], 2 + y + x * 5, 0)
                for z in range(7):
                    self.btns[self.i] = Button("/", self)
                    self.btns[self.i].setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
                    self.layout.addWidget(self.btns[self.i], y + 2 + x * 5, z + 1)
                    self.i += 1
                    
        self.spremi.clicked.connect(self.toCSV)
        self.uvezi.clicked.connect(self.uvoz)
        self.prebaci.clicked.connect(self.izvoz)
        self.layout.addWidget(self.spremi, 1 + (maximum - minimum + 1) * 5, 3)
        self.layout.addWidget(self.uvezi, 1 + (maximum - minimum + 1) * 5, 4)
        self.layout.addWidget(self.prebaci, 1 + (maximum - minimum + 1) * 5, 5)
        
        self.ubaci(SL, minimum, date, 2)
                
        self.ubaci(ST1, minimum, date, 3)
                
        self.ubaci(ST2, minimum, date, 4)
                
        self.ubaci(ostalo, minimum, date, 5)
                
        
        self.widget.setLayout(self.layout)
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(self.widget)

        self.setCentralWidget(self.scroll)

        self.setWindowTitle('Raspored')
        self.showMaximized()

    def dragEnterEvent(self, e):
        e.accept()
        
    def dragMoveEvent(self, e):
        source = e.source()
        target = QApplication.widgetAt(self.mapToGlobal(e.pos()))
        if (isinstance(e.source(), Button) and isinstance(target, Button) and target != source):
            e.accept()
        else:
            e.ignore()

    def dropEvent(self, e):
        source = e.source()
        target = QApplication.widgetAt(self.mapToGlobal(e.pos()))
        if (not isinstance(source, Button) or not isinstance(target, Button) 
            or target == source):
                return
        layout = self.widget.layout()

        sourceIndex = layout.indexOf(source)
        sourcePos = layout.getItemPosition(sourceIndex)

        targetIndex = layout.indexOf(target)
        targetPos = layout.getItemPosition(targetIndex)
        
        layout.addWidget(source, *sourcePos)
        layout.addWidget(target, *targetPos)
        
        tekst = source.text()
        source.setText(target.text())
        target.setText(tekst)

        e.accept()


def main():
    global app, main
    #app = QApplication(sys.argv)
    main = MainWindow()
    main.hide()
    #main.show()
    #app.exec_()


def showWindow():
    global main
    main.show()

def reopen():
    global main
    main.close()
    main = MainWindow()
#if __name__ == '__main__':
    #main()

